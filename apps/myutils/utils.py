# -*- coding: utf-8 -*-
__author__ = "dzt"
__date__ = "2019/12/24"
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import colors
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side
import math
import xlsxwriter

data_dict = [
    {
        "road_type": "违停",
        "road_code": "91001",
        "address": "大学城保税港",
        "owner_company": "沙区教委",
        "longitude_latitude": "106.123,29.123",
        "build_company": "重庆恒通",
        "device_code": "50000000000000001",
        "brigade": "五大队",
        "time": "2018年10月09号",
        "camera_list": [
            {
                "ip": "50.22.32.5:10000",
                "device_type": "TEC-300（海康）",
                "username_pwd": "admin/hik12345",
                "terminal": "50.22.32.5",
                "server": "50.22.32.248",
                "wf_type": "10395,13452,13453,13011,10482,12082,13011,13011,13011,13011"
            },
            {
                "ip": "50.22.32.5:12000",
                "device_type": "TEC-300（海康）",
                "username_pwd": "admin/hik12345",
                "terminal": "50.22.32.5",
                "server": "50.22.32.248",
                "wf_type": "10395,13453"
            },
        ],
        "terminal_list": [
            {"ip": "50.22.32.5", "device_type": "TEC-300（海康）", "username_pwd": "admin/hik12345", "terminal": "",
             "server": "50.22.32.248", "wf_type": ""},
        ],
        "other_list": [
            {
                "device_name": "电子警察悬臂杆",
                "device_type": "八棱H6.5M,L9-15M",
                "device_num": "3",
            },
            {
                "device_name": "室外挂箱",
                "device_type": "200*300*400",
                "device_num": "3",
            },
            {
                "device_name": "补光灯",
                "device_type": "CXBG-1-PS-DS-TL2002A",
                "device_num": "13",
            },
            {
                "device_name": "终端盒",
                "device_type": "4口",
                "device_num": "4",
            },
            {
                "device_name": "落地式机箱",
                "device_type": "600*600*1200",
                "device_num": "1",
            },
            {
                "device_name": "光收发器",
                "device_type": "SA-100",
                "device_num": "3",
            },
            {
                "device_name": "网线",
                "device_type": "3米",
                "device_num": "3",
            }
        ],
        "map_img": 'G:\\dzt\\devices_mgmt\\media\\road\img\\111.png',
        "img2": 'G:\\dzt\\devices_mgmt\\media\\road\img\\222.jpg',
    },
    {
        "road_type": "违停2",
        "road_code": "91002",
        "address": "大学城保税港",
        "owner_company": "沙区教委",
        "longitude_latitude": "106.123,29.123",
        "build_company": "重庆恒通",
        "device_code": "50000000000000001",
        "brigade": "五大队",
        "time": "2018年10月09号",
        "camera_list": [
            {
                "ip": "50.22.32.5:10000",
                "device_type": "TEC-300（海康）",
                "username_pwd": "admin/hik12345",
                "terminal": "50.22.32.5",
                "server": "50.22.32.248",
                "wf_type": "10395,13452,13453,13011,10482,12082,13011,13011,13011,13011"
            },
            {
                "ip": "50.22.32.5:12000",
                "device_type": "TEC-300（海康）",
                "username_pwd": "admin/hik12345",
                "terminal": "50.22.32.5",
                "server": "50.22.32.248",
                "wf_type": "10395,13453"
            },
        ],
        "terminal_list": [
            {
                "ip": "50.22.32.5",
                "device_type": "TEC-300（海康）",
                "username_pwd": "admin/hik12345",
                "terminal": "",
                "server": "50.22.32.248", "wf_type": ""},
        ],
        "other_list": [
            {
                "device_name": "电子警察悬臂杆",
                "device_type": "八棱H6.5M,L9-15M",
                "device_num": "3",
            },
            {
                "device_name": "室外挂箱",
                "device_type": "200*300*400",
                "device_num": "3",
            },
            {
                "device_name": "补光灯",
                "device_type": "CXBG-1-PS-DS-TL2002A",
                "device_num": "13",
            },
            {
                "device_name": "终端盒",
                "device_type": "4口",
                "device_num": "4",
            },
            {
                "device_name": "落地式机箱",
                "device_type": "600*600*1200",
                "device_num": "1",
            },
            {
                "device_name": "光收发器",
                "device_type": "SA-100",
                "device_num": "3",
            },
            {
                "device_name": "网线",
                "device_type": "3米",
                "device_num": "3",
            }
        ],
        "map_img": 'G:\\dzt\\devices_mgmt\\media\\road\img\\111.png',
        "img2": 'G:\\dzt\\devices_mgmt\\media\\road\img\\222.jpg',
    },
]


def export_excel(data_dict, save_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "封面"
    for data in data_dict:
        # Font 字体
        # TODO
        ws = wb.create_sheet(data['road_code'])
        ws.title = data['road_code']

        FontStyle = Font(name='宋体', size=11, bold=True, color=colors.BLACK)
        FontStyle2 = Font(name='宋体', size=11, color=colors.BLACK)
        FontStyle3 = Font(name='宋体', size=11, bold=True, color=colors.RED)
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))

        ws['A1'].font = FontStyle
        ws['A1'] = '设备信息'
        ws.row_dimensions[1].height = 24
        ws['A1'].alignment = align  # 设置居中对齐后再合并
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 24
        ws.merge_cells('A1:B1')

        ws['C1'].font = FontStyle
        ws['C1'] = '安装点位信息'
        ws['C1'].alignment = align
        ws.column_dimensions['C'].width = 24
        ws.column_dimensions['D'].width = 24
        ws.column_dimensions['E'].width = 24
        ws.merge_cells('C1:E1')

        ws['F1'].font = FontStyle
        ws['F1'] = '建设信息'
        ws['F1'].alignment = align
        ws.column_dimensions['F'].width = 24
        ws.column_dimensions['G'].width = 24
        ws.merge_cells('F1:G1')

        ws['A2'].font = FontStyle2
        ws['A2'] = '设备类型'
        ws.row_dimensions[2].height = 24
        ws['A2'].alignment = align
        # TODO
        ws['B2'].font = FontStyle2
        ws['B2'] = data['road_type']
        ws.row_dimensions[2].height = 24
        ws['B2'].alignment = align
        ws['C2'].font = FontStyle2
        ws['C2'] = '安装地点名称'
        ws.row_dimensions[2].height = 24
        ws['C2'].alignment = align
        # TODO
        ws['D2'].font = FontStyle2
        ws['D2'] = data['address']
        ws.row_dimensions[2].height = 24
        ws['D2'].alignment = align
        ws.merge_cells('D2:E2')

        ws['F2'].font = FontStyle2
        ws['F2'] = '业主单位'
        ws.row_dimensions[2].height = 24
        ws['F2'].alignment = align
        # TODO
        ws['G2'].font = FontStyle2
        ws['G2'] = data['owner_company']
        ws.row_dimensions[2].height = 24
        ws['G2'].alignment = align

        ws['A3'].font = FontStyle3
        ws['A3'] = '设备档案编号'
        ws.row_dimensions[3].height = 24
        ws['A3'].alignment = align
        # TODO
        ws['B3'].font = FontStyle3
        ws['B3'] = data['road_code']
        ws.row_dimensions[3].height = 24
        ws['B3'].alignment = align
        ws['C3'].font = FontStyle2
        ws['C3'] = '坐标（经纬度）'
        ws.row_dimensions[3].height = 24
        ws['C3'].alignment = align
        # TODO
        ws['D3'].font = FontStyle2
        ws['D3'] = data['longitude_latitude']
        ws.row_dimensions[3].height = 24
        ws['D3'].alignment = align
        ws.merge_cells('D3:E3')

        ws['F3'].font = FontStyle2
        ws['F3'] = '建设单位'
        ws.row_dimensions[3].height = 24
        ws['F3'].alignment = align
        # TODO
        ws['G3'].font = FontStyle2
        ws['G3'] = data['build_company']
        ws.row_dimensions[3].height = 24
        ws['G3'].alignment = align

        ws['A4'].font = FontStyle2
        ws['A4'] = '执法设备编号'
        ws.row_dimensions[4].height = 24
        ws['A4'].alignment = align
        # TODO
        ws['B4'].font = FontStyle2
        ws['B4'] = data['device_code']
        ws.row_dimensions[4].height = 24
        ws['B4'].alignment = align
        ws['C4'].font = FontStyle2
        ws['C4'] = '所属大队'
        ws.row_dimensions[4].height = 24
        ws['C4'].alignment = align
        # TODO
        ws['D4'].font = FontStyle2
        ws['D4'] = data['brigade']
        ws.row_dimensions[4].height = 24
        ws['D4'].alignment = align
        ws.merge_cells('D4:E4')
        ws['F4'].font = FontStyle2
        ws['F4'] = '建设移交时间'
        ws.row_dimensions[4].height = 24
        ws['F4'].alignment = align
        # TODO
        ws['G4'].font = FontStyle2
        ws['G4'] = data['time']
        ws.row_dimensions[4].height = 24
        ws['G4'].alignment = align

        ws['A5'].font = FontStyle
        ws['A5'] = '电子警察点位主要设备信息'
        ws.row_dimensions[5].height = 24
        ws['A5'].alignment = align  # 设置居中对齐后再合并
        ws.merge_cells('A5:G5')

        ws['A6'].font = FontStyle2
        ws['A6'] = '设备名称'
        ws.row_dimensions[6].height = 24
        ws['A6'].alignment = align
        ws['B6'].font = FontStyle2
        ws['B6'] = 'IP地址'
        ws.row_dimensions[6].height = 24
        ws['B6'].alignment = align
        ws['C6'].font = FontStyle2
        ws['C6'] = '规格型号（厂家）'
        ws.row_dimensions[6].height = 24
        ws['C6'].alignment = align
        ws['D6'].font = FontStyle2
        ws['D6'] = '用户名/密码'
        ws.row_dimensions[6].height = 24
        ws['D6'].alignment = align
        ws['E6'].font = FontStyle2
        ws['E6'] = '接入终端地址'
        ws.row_dimensions[6].height = 24
        ws['E6'].alignment = align
        ws['F6'].font = FontStyle2
        ws['F6'] = '接入服务器地址'
        ws.row_dimensions[6].height = 24
        ws['F6'].alignment = align
        ws['G6'].font = FontStyle2
        ws['G6'] = '违法代码'
        ws.row_dimensions[6].height = 24
        ws['G6'].alignment = align

        camera_list = data['camera_list']

        terminal_list = data['terminal_list']

        index = 7
        for i in camera_list:
            ws_index = "A" + str(index)

            ws[ws_index].font = FontStyle2
            ws[ws_index] = "抓拍摄像机"
            ws[ws_index].alignment = align
            ws.row_dimensions[index].height = 40

            ws_index = "B" + str(index)

            ws[ws_index].font = FontStyle2
            ws[ws_index] = i['ip']
            ws[ws_index].alignment = align

            ws_index = "C" + str(index)

            ws[ws_index].font = FontStyle2
            ws[ws_index] = i['device_type']
            ws[ws_index].alignment = align

            ws_index = "D" + str(index)

            ws[ws_index].font = FontStyle2
            ws[ws_index] = i['username_pwd']
            ws[ws_index].alignment = align

            ws_index = "E" + str(index)
            ws[ws_index].font = FontStyle2
            ws[ws_index] = i['terminal']
            ws[ws_index].alignment = align

            ws_index = "F" + str(index)
            ws[ws_index].font = FontStyle2
            ws[ws_index] = i['server']
            ws[ws_index].alignment = align

            ws_index = "G" + str(index)
            ws[ws_index].font = FontStyle2
            ws[ws_index] = i['wf_type']
            ws[ws_index].alignment = align

            index += 1

        for i in terminal_list:
            ws_index = "A" + str(index)

            ws[ws_index].font = FontStyle2
            ws[ws_index] = "终端服务器"
            ws[ws_index].alignment = align
            ws.row_dimensions[index].height = 24

            ws_index = "B" + str(index)

            ws[ws_index].font = FontStyle2
            ws[ws_index] = i['ip']
            ws[ws_index].alignment = align

            ws_index = "C" + str(index)

            ws[ws_index].font = FontStyle2
            ws[ws_index] = i['device_type']
            ws[ws_index].alignment = align

            ws_index = "D" + str(index)

            ws[ws_index].font = FontStyle2
            ws[ws_index] = i['username_pwd']
            ws[ws_index].alignment = align

            ws_index = "E" + str(index)
            ws[ws_index].font = FontStyle2
            ws[ws_index] = i['terminal']
            ws[ws_index].alignment = align

            ws_index = "F" + str(index)
            ws[ws_index].font = FontStyle2
            ws[ws_index] = i['server']
            ws[ws_index].alignment = align

            ws_index = "G" + str(index)
            ws[ws_index].font = FontStyle2
            ws[ws_index] = i['wf_type']
            ws[ws_index].alignment = align

            index += 1

        ws_index = "A" + str(index)
        ws_index2 = "G" + str(index)
        ws[ws_index].font = FontStyle
        ws[ws_index] = '抓拍点位主要杆件及材料信息'
        ws.row_dimensions[index].height = 24
        ws[ws_index].alignment = align  # 设置居中对齐后再合并
        ws.merge_cells('%s:%s' % (ws_index, ws_index2))
        index += 1

        ws_index = "A" + str(index)
        ws[ws_index].font = FontStyle2
        ws[ws_index] = '名称'
        ws.row_dimensions[index].height = 24
        ws[ws_index].alignment = align

        ws_index = "B" + str(index)
        ws[ws_index].font = FontStyle2
        ws[ws_index] = '规格'
        ws[ws_index].alignment = align

        ws_index = "C" + str(index)
        ws[ws_index].font = FontStyle2
        ws[ws_index] = '数量'
        ws[ws_index].alignment = align

        ws_index = "D" + str(index)
        ws[ws_index].font = FontStyle2
        ws[ws_index] = ''
        ws[ws_index].alignment = align

        ws_index = "E" + str(index)
        ws[ws_index].font = FontStyle2
        ws[ws_index] = '名称'
        ws[ws_index].alignment = align

        ws_index = "F" + str(index)
        ws[ws_index].font = FontStyle2
        ws[ws_index] = '规格'
        ws[ws_index].alignment = align

        ws_index = "G" + str(index)
        ws[ws_index].font = FontStyle2
        ws[ws_index] = '数量'
        ws[ws_index].alignment = align

        other_list = data['other_list']

        other_list = [other_list[i:i + 2] for i in range(0, len(other_list), 2)]
        for i in other_list:

            if len(i) == 1:
                x = i[0]
                ws_index = "A" + str(index)
                ws[ws_index].font = FontStyle2
                ws[ws_index] = x['device_name']
                ws.row_dimensions[index].height = 24
                ws[ws_index].alignment = align

                ws_index = "B" + str(index)
                ws[ws_index].font = FontStyle2
                ws[ws_index] = x['device_type']
                ws[ws_index].alignment = align

                ws_index = "C" + str(index)
                ws[ws_index].font = FontStyle2
                ws[ws_index] = x['device_num']
                ws[ws_index].alignment = align
            elif len(i) == 2:
                x = i[0]
                x1 = i[1]
                ws_index = "A" + str(index)
                ws[ws_index].font = FontStyle2
                ws[ws_index] = x['device_name']
                ws.row_dimensions[index].height = 24
                ws[ws_index].alignment = align

                ws_index = "B" + str(index)
                ws[ws_index].font = FontStyle2
                ws[ws_index] = x['device_type']
                ws[ws_index].alignment = align

                ws_index = "C" + str(index)
                ws[ws_index].font = FontStyle2
                ws[ws_index] = x['device_num']
                ws[ws_index].alignment = align

                ws_index = "E" + str(index)
                ws[ws_index].font = FontStyle2
                ws[ws_index] = x['device_name']
                ws[ws_index].alignment = align

                ws_index = "F" + str(index)
                ws[ws_index].font = FontStyle2
                ws[ws_index] = x['device_type']
                ws[ws_index].alignment = align

                ws_index = "G" + str(index)
                ws[ws_index].font = FontStyle2
                ws[ws_index] = x['device_num']
                ws[ws_index].alignment = align

            index += 1

        ws_index = "A" + str(index)
        ws_index2 = "G" + str(index)
        ws[ws_index].font = FontStyle
        ws[ws_index] = '其他信息'
        ws.row_dimensions[index].height = 24
        ws[ws_index].alignment = align  # 设置居中对齐后再合并
        ws.merge_cells('%s:%s' % (ws_index, ws_index2))
        index += 1

        other = "电源：引入路口路灯配电箱，光纤网络联通专网"
        ws_index = "A" + str(index)
        ws_index2 = "G" + str(index)
        ws[ws_index].font = FontStyle
        ws[ws_index] = other
        ws.row_dimensions[index].height = 24
        ws[ws_index].alignment = align  # 设置居中对齐后再合并
        ws.merge_cells('%s:%s' % (ws_index, ws_index2))
        index += 1

        ws_index = "A" + str(index)
        ws_index2 = "C" + str(index)
        ws_index3 = "D" + str(index)
        ws_index4 = "G" + str(index)
        ws[ws_index].font = FontStyle
        ws[ws_index3].font = FontStyle
        ws[ws_index] = '点位示意图'
        ws[ws_index3] = '现场照片图'
        ws.row_dimensions[index].height = 24
        ws[ws_index].alignment = align  # 设置居中对齐后再合并
        ws[ws_index3].alignment = align  # 设置居中对齐后再合并
        ws.merge_cells('%s:%s' % (ws_index, ws_index2))
        ws.merge_cells('%s:%s' % (ws_index3, ws_index4))
        index += 1

        map_img = data['map_img']
        img2 = data['img2']

        ws_index = "A" + str(index)
        ws_index2 = "D" + str(index)

        try:
            img = Image(map_img)
            newsize = (540, 400)
            img.width, img.height = newsize
            ws.add_image(img, ws_index)

            img = Image(img2)
            newsize = (770, 400)
            img.width, img.height = newsize
            ws.add_image(img, ws_index2)
        except Exception as e:
            print(e)

        # 添加边框
        for i in range(1, 40):
            s = str(i)
            ws_index1 = "A" + s
            ws_index2 = "B" + s
            ws_index3 = "C" + s
            ws_index4 = "D" + s
            ws_index5 = "E" + s
            ws_index6 = "F" + s
            ws_index7 = "G" + s
            ws[ws_index1].border = border
            ws[ws_index2].border = border
            ws[ws_index3].border = border
            ws[ws_index4].border = border
            ws[ws_index5].border = border
            ws[ws_index6].border = border
            ws[ws_index7].border = border

    # wb.save("test.xlsx")
    wb.save(save_file)


# 坐标转换

x_pi = 3.14159265358979324 * 3000.0 / 180.0
pi = 3.1415926535897932384626  # π
a = 6378245.0  # 长半轴
ee = 0.00669342162296594323  # 偏心率平方


def bd09_to_gcj02(bd_lon, bd_lat):
    """
    百度坐标系(BD-09)转火星坐标系(GCJ-02)
    百度——>谷歌、高德
    :param bd_lat:百度坐标纬度
    :param bd_lon:百度坐标经度
    :return:转换后的坐标列表形式
    """
    x = bd_lon - 0.0065
    y = bd_lat - 0.006
    z = math.sqrt(x * x + y * y) - 0.00002 * math.sin(y * x_pi)
    theta = math.atan2(y, x) - 0.000003 * math.cos(x * x_pi)
    gg_lng = z * math.cos(theta)
    gg_lat = z * math.sin(theta)
    return [gg_lng, gg_lat]


# 导出台账
def export_2_excel(all_dj_datas, sheet_info, filename):
    # print(all_wt_datas)

    title = ['路口名称', '设备类型', '执法设备编号', '相机', '终端', '建设单位', '经纬度', '是否有效', '点位示意图', '现场照片']

    f = xlsxwriter.Workbook(filename)
    fsheet = f.add_worksheet(sheet_info)
    # 首行格式
    format1 = f.add_format({
        'bold': True, 'font_color': 'black', 'font_size': 15, 'align': 'center', 'font_name': '宋体'
    })
    # 内容格式
    format2 = f.add_format({'font_color': 'black', 'font_size': 12, 'align': 'center', 'font_name': '宋体'})
    format3 = f.add_format({'font_color': 'blue', 'font_size': 12, 'align': 'center', 'font_name': '宋体'})

    # 设置A-K列 宽 20
    fsheet.set_column("A:L", 20)
    # 写入首行
    for t in range(len(title)):
        fsheet.write(0, t, title[t], format1)
    # 写入数据
    for i in range(len(all_dj_datas)):
        wt_data = all_dj_datas[i]
        for x in range(len(wt_data)):
            # if wt_data[x] and x > 9:
            #     fsheet.write_url(i + 1, x, wt_data[x], format3, string='点击查看')
            # else:
            #     fsheet.write(i + 1, x, wt_data[x], format2)
            fsheet.write(i + 1, x, wt_data[x], format2)
    f.close()


if __name__ == '__main__':
    export_excel(data_dict, "test.xlsx")
