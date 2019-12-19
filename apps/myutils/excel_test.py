# -*- coding: utf-8 -*-
__author__ = "dzt"
__date__ = "2019/12/19"
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import colors
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side

wb = openpyxl.Workbook()

# Font 字体
ws = wb.active
# TODO
ws.title = '91001'

# default 11pt, Calibri
# italic24Font = Font(size=24, italic=True)  # 24号斜体
# Font24 = Font(size=24)
# ws['A2'].font = Font24
# ws['A2'] = '24号'

FontStyle = Font(name='宋体', size=11, bold=True, color=colors.BLACK)
FontStyle2 = Font(name='宋体', size=11, color=colors.BLACK)
FontStyle3 = Font(name='宋体', size=11, bold=True, color=colors.RED)
align = Alignment(horizontal='center', vertical='center', wrap_text=True)
border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))

ws['A1'].font = FontStyle
ws['A1'] = '设备信息'
ws.row_dimensions[1].height = 24
ws['A1'].alignment = align  # 设置居中对齐后再合并
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 24
ws.merge_cells('A1:B1')


ws['C1'].font = FontStyle
ws['C1'] = '安装点位信息'
ws['C1'].alignment = align
ws.column_dimensions['C'].width = 24
ws.column_dimensions['D'].width = 24
ws.column_dimensions['E'].width = 24
ws.merge_cells('C1:E1')

ws['F1'].font = FontStyle
ws['F1'] = '建设信息'
ws['F1'].alignment = align
ws.column_dimensions['F'].width = 24
ws.column_dimensions['G'].width = 24
ws.merge_cells('F1:G1')

ws['A2'].font = FontStyle2
ws['A2'] = '设备类型'
ws.row_dimensions[2].height = 24
ws['A2'].alignment = align
# TODO
ws['B2'].font = FontStyle2
ws['B2'] = '违停'
ws.row_dimensions[2].height = 24
ws['B2'].alignment = align
ws['C2'].font = FontStyle2
ws['C2'] = '安装地点名称'
ws.row_dimensions[2].height = 24
ws['C2'].alignment = align
# TODO
ws['D2'].font = FontStyle2
ws['D2'] = '大学城保税港'
ws.row_dimensions[2].height = 24
ws['D2'].alignment = align
ws.merge_cells('D2:E2')

ws['F2'].font = FontStyle2
ws['F2'] = '业主单位'
ws.row_dimensions[2].height = 24
ws['F2'].alignment = align
# TODO
ws['G2'].font = FontStyle2
ws['G2'] = '沙区教委'
ws.row_dimensions[2].height = 24
ws['G2'].alignment = align

ws['A3'].font = FontStyle3
ws['A3'] = '设备档案编号'
ws.row_dimensions[3].height = 24
ws['A3'].alignment = align
# TODO
ws['B3'].font = FontStyle3
ws['B3'] = '91001'
ws.row_dimensions[3].height = 24
ws['B3'].alignment = align
ws['C3'].font = FontStyle2
ws['C3'] = '坐标（经纬度）'
ws.row_dimensions[3].height = 24
ws['C3'].alignment = align
# TODO
ws['D3'].font = FontStyle2
ws['D3'] = '106.123,29.123'
ws.row_dimensions[3].height = 24
ws['D3'].alignment = align
ws.merge_cells('D3:E3')

ws['F3'].font = FontStyle2
ws['F3'] = '建设单位'
ws.row_dimensions[3].height = 24
ws['F3'].alignment = align
# TODO
ws['G3'].font = FontStyle2
ws['G3'] = '重庆恒通'
ws.row_dimensions[3].height = 24
ws['G3'].alignment = align

ws['A4'].font = FontStyle2
ws['A4'] = '执法设备编号'
ws.row_dimensions[4].height = 24
ws['A4'].alignment = align
# TODO
ws['B4'].font = FontStyle2
ws['B4'] = '50000000000000001'
ws.row_dimensions[4].height = 24
ws['B4'].alignment = align
ws['C4'].font = FontStyle2
ws['C4'] = '所属大队'
ws.row_dimensions[4].height = 24
ws['C4'].alignment = align
# TODO
ws['D4'].font = FontStyle2
ws['D4'] = '五大队'
ws.row_dimensions[4].height = 24
ws['D4'].alignment = align
ws.merge_cells('D4:E4')
ws['F4'].font = FontStyle2
ws['F4'] = '建设移交时间'
ws.row_dimensions[4].height = 24
ws['F4'].alignment = align
# TODO
ws['G4'].font = FontStyle2
ws['G4'] = '2018年10月09号'
ws.row_dimensions[4].height = 24
ws['G4'].alignment = align

ws['A5'].font = FontStyle
ws['A5'] = '电子警察点位主要设备信息'
ws.row_dimensions[5].height = 24
ws['A5'].alignment = align  # 设置居中对齐后再合并
ws.merge_cells('A5:G5')

ws['A6'].font = FontStyle2
ws['A6'] = '设备名称'
ws.row_dimensions[6].height = 24
ws['A6'].alignment = align
ws['B6'].font = FontStyle2
ws['B6'] = 'IP地址'
ws.row_dimensions[6].height = 24
ws['B6'].alignment = align
ws['C6'].font = FontStyle2
ws['C6'] = '规格型号（厂家）'
ws.row_dimensions[6].height = 24
ws['C6'].alignment = align
ws['D6'].font = FontStyle2
ws['D6'] = '用户名/密码'
ws.row_dimensions[6].height = 24
ws['D6'].alignment = align
ws['E6'].font = FontStyle2
ws['E6'] = '接入终端地址'
ws.row_dimensions[6].height = 24
ws['E6'].alignment = align
ws['F6'].font = FontStyle2
ws['F6'] = '接入服务器地址'
ws.row_dimensions[6].height = 24
ws['F6'].alignment = align
ws['G6'].font = FontStyle2
ws['G6'] = '违法代码'
ws.row_dimensions[6].height = 24
ws['G6'].alignment = align

camera_list = [
    {
        "ip": "50.22.32.5:10000",
        "device_type": "TEC-300（海康）",
        "username_pwd": "admin/hik12345",
        "terminal": "50.22.32.5",
        "server": "50.22.32.248",
        "wf_type": "10395,13452,13453,13011,10482,12082,13011,13011,13011,13011"
    },
    {
        "ip": "50.22.32.5:12000",
        "device_type": "TEC-300（海康）",
        "username_pwd": "admin/hik12345",
        "terminal": "50.22.32.5",
        "server": "50.22.32.248",
        "wf_type": "10395,13453"
    },
]

terminal_list = [
    {"ip": "50.22.32.5", "device_type": "TEC-300（海康）", "username_pwd": "admin/hik12345", "terminal": "",
     "server": "50.22.32.248", "wf_type": ""},
]

index = 7
for i in camera_list:
    ws_index = "A" + str(index)

    ws[ws_index].font = FontStyle2
    ws[ws_index] = "抓拍摄像机"
    ws[ws_index].alignment = align
    ws.row_dimensions[index].height = 40

    ws_index = "B" + str(index)

    ws[ws_index].font = FontStyle2
    ws[ws_index] = i['ip']
    ws[ws_index].alignment = align

    ws_index = "C" + str(index)

    ws[ws_index].font = FontStyle2
    ws[ws_index] = i['device_type']
    ws[ws_index].alignment = align

    ws_index = "D" + str(index)

    ws[ws_index].font = FontStyle2
    ws[ws_index] = i['username_pwd']
    ws[ws_index].alignment = align

    ws_index = "E" + str(index)
    ws[ws_index].font = FontStyle2
    ws[ws_index] = i['terminal']
    ws[ws_index].alignment = align

    ws_index = "F" + str(index)
    ws[ws_index].font = FontStyle2
    ws[ws_index] = i['server']
    ws[ws_index].alignment = align

    ws_index = "G" + str(index)
    ws[ws_index].font = FontStyle2
    ws[ws_index] = i['wf_type']
    ws[ws_index].alignment = align

    index += 1

for i in terminal_list:
    ws_index = "A" + str(index)

    ws[ws_index].font = FontStyle2
    ws[ws_index] = "终端服务器"
    ws[ws_index].alignment = align
    ws.row_dimensions[index].height = 24

    ws_index = "B" + str(index)

    ws[ws_index].font = FontStyle2
    ws[ws_index] = i['ip']
    ws[ws_index].alignment = align

    ws_index = "C" + str(index)

    ws[ws_index].font = FontStyle2
    ws[ws_index] = i['device_type']
    ws[ws_index].alignment = align

    ws_index = "D" + str(index)

    ws[ws_index].font = FontStyle2
    ws[ws_index] = i['username_pwd']
    ws[ws_index].alignment = align

    ws_index = "E" + str(index)
    ws[ws_index].font = FontStyle2
    ws[ws_index] = i['terminal']
    ws[ws_index].alignment = align

    ws_index = "F" + str(index)
    ws[ws_index].font = FontStyle2
    ws[ws_index] = i['server']
    ws[ws_index].alignment = align

    ws_index = "G" + str(index)
    ws[ws_index].font = FontStyle2
    ws[ws_index] = i['wf_type']
    ws[ws_index].alignment = align

    index += 1

ws_index = "A" + str(index)
ws_index2 = "G" + str(index)
ws[ws_index].font = FontStyle
ws[ws_index] = '抓拍点位主要杆件及材料信息'
ws.row_dimensions[index].height = 24
ws[ws_index].alignment = align  # 设置居中对齐后再合并
ws.merge_cells('%s:%s' % (ws_index, ws_index2))
index += 1

ws_index = "A" + str(index)
ws[ws_index].font = FontStyle2
ws[ws_index] = '名称'
ws.row_dimensions[index].height = 24
ws[ws_index].alignment = align

ws_index = "B" + str(index)
ws[ws_index].font = FontStyle2
ws[ws_index] = '规格'
ws[ws_index].alignment = align

ws_index = "C" + str(index)
ws[ws_index].font = FontStyle2
ws[ws_index] = '数量'
ws[ws_index].alignment = align

ws_index = "D" + str(index)
ws[ws_index].font = FontStyle2
ws[ws_index] = ''
ws[ws_index].alignment = align

ws_index = "E" + str(index)
ws[ws_index].font = FontStyle2
ws[ws_index] = '名称'
ws[ws_index].alignment = align

ws_index = "F" + str(index)
ws[ws_index].font = FontStyle2
ws[ws_index] = '规格'
ws[ws_index].alignment = align

ws_index = "G" + str(index)
ws[ws_index].font = FontStyle2
ws[ws_index] = '数量'
ws[ws_index].alignment = align

other_list = [
    {
        "device_name": "电子警察悬臂杆",
        "device_type": "八棱H6.5M,L9-15M",
        "device_num": "3",
    },
    {
        "device_name": "室外挂箱",
        "device_type": "200*300*400",
        "device_num": "3",
    },
    {
        "device_name": "补光灯",
        "device_type": "CXBG-1-PS-DS-TL2002A",
        "device_num": "13",
    },
    {
        "device_name": "终端盒",
        "device_type": "4口",
        "device_num": "4",
    },
    {
        "device_name": "落地式机箱",
        "device_type": "600*600*1200",
        "device_num": "1",
    },
    {
        "device_name": "光收发器",
        "device_type": "SA-100",
        "device_num": "3",
    },
    {
        "device_name": "网线",
        "device_type": "3米",
        "device_num": "3",
    }
]

other_list = [other_list[i:i + 2] for i in range(0, len(other_list), 2)]
for i in other_list:

    if len(i) == 1:
        x = i[0]
        ws_index = "A" + str(index)
        ws[ws_index].font = FontStyle2
        ws[ws_index] = x['device_name']
        ws.row_dimensions[index].height = 24
        ws[ws_index].alignment = align

        ws_index = "B" + str(index)
        ws[ws_index].font = FontStyle2
        ws[ws_index] = x['device_type']
        ws[ws_index].alignment = align

        ws_index = "C" + str(index)
        ws[ws_index].font = FontStyle2
        ws[ws_index] = x['device_num']
        ws[ws_index].alignment = align
    elif len(i) == 2:
        x = i[0]
        x1 = i[1]
        ws_index = "A" + str(index)
        ws[ws_index].font = FontStyle2
        ws[ws_index] = x['device_name']
        ws.row_dimensions[index].height = 24
        ws[ws_index].alignment = align

        ws_index = "B" + str(index)
        ws[ws_index].font = FontStyle2
        ws[ws_index] = x['device_type']
        ws[ws_index].alignment = align

        ws_index = "C" + str(index)
        ws[ws_index].font = FontStyle2
        ws[ws_index] = x['device_num']
        ws[ws_index].alignment = align

        ws_index = "E" + str(index)
        ws[ws_index].font = FontStyle2
        ws[ws_index] = x['device_name']
        ws[ws_index].alignment = align

        ws_index = "F" + str(index)
        ws[ws_index].font = FontStyle2
        ws[ws_index] = x['device_type']
        ws[ws_index].alignment = align

        ws_index = "G" + str(index)
        ws[ws_index].font = FontStyle2
        ws[ws_index] = x['device_num']
        ws[ws_index].alignment = align

    index += 1

ws_index = "A" + str(index)
ws_index2 = "G" + str(index)
ws[ws_index].font = FontStyle
ws[ws_index] = '其他信息'
ws.row_dimensions[index].height = 24
ws[ws_index].alignment = align  # 设置居中对齐后再合并
ws.merge_cells('%s:%s' % (ws_index, ws_index2))
index += 1

other = "电源：引入路口路灯配电箱，光纤网络联通专网"
ws_index = "A" + str(index)
ws_index2 = "G" + str(index)
ws[ws_index].font = FontStyle
ws[ws_index] = other
ws.row_dimensions[index].height = 24
ws[ws_index].alignment = align  # 设置居中对齐后再合并
ws.merge_cells('%s:%s' % (ws_index, ws_index2))
index += 1

ws_index = "A" + str(index)
ws_index2 = "C" + str(index)
ws_index3 = "D" + str(index)
ws_index4 = "G" + str(index)
ws[ws_index].font = FontStyle
ws[ws_index3].font = FontStyle
ws[ws_index] = '点位示意图'
ws[ws_index3] = '现场照片图'
ws.row_dimensions[index].height = 24
ws[ws_index].alignment = align  # 设置居中对齐后再合并
ws[ws_index3].alignment = align  # 设置居中对齐后再合并
ws.merge_cells('%s:%s' % (ws_index, ws_index2))
ws.merge_cells('%s:%s' % (ws_index3, ws_index4))
index += 1

map_img = 'G:\\dzt\\devices_mgmt\\media\\road\img\\111.png'
img2 = 'G:\\dzt\\devices_mgmt\\media\\road\img\\222.jpg'

ws_index = "A" + str(index)
ws_index2 = "D" + str(index)

img = Image(map_img)
newsize = (540, 400)
img.width, img.height = newsize
ws.add_image(img, ws_index)

img = Image(img2)
newsize = (770, 400)
img.width, img.height = newsize
ws.add_image(img, ws_index2)

# 添加边框
for i in range(1, 40):
    s = str(i)
    ws_index1 = "A" + s
    ws_index2 = "B" + s
    ws_index3 = "C" + s
    ws_index4 = "D" + s
    ws_index5 = "E" + s
    ws_index6 = "F" + s
    ws_index7 = "G" + s
    ws[ws_index1].border = border
    ws[ws_index2].border = border
    ws[ws_index3].border = border
    ws[ws_index4].border = border
    ws[ws_index5].border = border
    ws[ws_index6].border = border
    ws[ws_index7].border = border

# # Unmerging cells 拆分单元格
# ws = wb.copy_worksheet(wb.get_sheet_by_name('merged'))  # 拷贝一份
# ws.title = 'unmerged'
# ws.unmerge_cells('A1:D3')
# ws.unmerge_cells('C5:D5')


wb.save('style2.xlsx')
