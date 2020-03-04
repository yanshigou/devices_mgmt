__author__ = 'dzt'
__date__ = '2019/12/14 0:29'

import xlrd
import xlwt
import re
import openpyxl
from datetime import datetime
import requests
import json


file = 'C:\\Users\\Administrator\\Desktop\\banan.xlsx'


def read_excel():
    wb = xlrd.open_workbook(filename=file)  # 打开文件
    # print(wb.sheet_names())  # 获取所有表格名字

    sheet = wb.sheet_by_index(0)  # 通过索引获取表格
    print(sheet.name, sheet.nrows, sheet.ncols)

    # outwb = openpyxl.Workbook()  # 打开一个将写的文件
    # outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    #
    ok = 0
    fail = 0
    for i in range(3, sheet.nrows):
        rows = sheet.row_values(i)  # 获取行内容
        print(rows)
        d_type = "wt"
        camera_ip = rows[2]
        terminal_ip = rows[3]
        password = rows[4]
        device_code = rows[5]
        road_code = device_code[-5:]
        address = rows[6]
        longitude = rows[7]
        latitude = rows[8]
        device_company = rows[9]
        if "海康" in device_company:
            device_company = 'hk'
        elif device_company == "科达":
            device_company = 'kd'
        elif device_company == "大华":
            device_company = 'dh'
        else:
            device_company = "未知"
        device_type = rows[10]
        server = rows[11]
        if server == "50.44.14.12":
            server = "1"
        elif server == "50.44.14.9":
            server = "2"
        elif server == "50.44.14.3":
            server = "3"
        else:
            server = None
        time = rows[18]
        if time:
            try:
                time = datetime.strptime(str(int(time)), "%Y%m%d")
                time = datetime.strftime(time, "%Y-%m-%d")
            except Exception as e:
                print(e)
                time = datetime.strptime(str(int(time)), "%Y%m")
                time = datetime.strftime(time, "%Y-%m-%d")
        else:
            time = None
        other = rows[13]
        owner_company = rows[14]
        if owner_company == "交巡警支队":
            owner_company = "1"
        elif owner_company == "木洞镇":
            owner_company = "2"
        elif owner_company == "界石镇":
            owner_company = "3"
        elif owner_company == "东泉镇":
            owner_company = "4"
        elif owner_company == "惠民镇":
            owner_company = "5"
        elif owner_company == "一品镇":
            owner_company = "6"
        elif owner_company == "融汇地产":
            owner_company = "7"
        elif owner_company == "南泉镇":
            owner_company = "8"
        elif owner_company == "姜家镇":
            owner_company = "9"
        elif owner_company == "风盛镇":
            owner_company = "10"
        elif owner_company == "安澜镇":
            owner_company = "11"
        elif owner_company == "重庆科锐":
            owner_company = "12"
        elif owner_company == "接龙镇":
            owner_company = "13"
        elif owner_company == "圣灯山镇":
            owner_company = "14"
        elif owner_company == "盛世江南地产":
            owner_company = "15"
        elif owner_company == "南彭镇":
            owner_company = "16"
        elif owner_company == "鱼洞街道":
            owner_company = "17"
        elif owner_company == "二圣镇":
            owner_company = "18"
        elif owner_company == "石滩镇":
            owner_company = "19"
        elif owner_company == "花溪街道":
            owner_company = "20"
        elif owner_company == "南泉街道":
            owner_company = "21"
        elif owner_company == "界石镇卫生院":
            owner_company = "22"
        else:
            owner_company = None

        build_company = rows[15]
        if build_company == "重庆企裕":
            build_company = "1"
        elif build_company == "重庆金马":
            build_company = "2"
        elif build_company == "重庆永惠安防":
            build_company = "3"
        elif build_company == "重庆社安科技":
            build_company = "4"
        elif build_company == "重庆莹轩":
            build_company = "5"
        elif build_company == "青岛海信":
            build_company = "6"
        elif build_company == "重庆渝达":
            build_company = "7"
        elif build_company == "重庆佳讯":
            build_company = "8"
        elif build_company == "重庆八达通":
            build_company = "9"
        elif build_company == "重庆宏昂":
            build_company = "10"
        else:
            build_company = None

        print(d_type, camera_ip, terminal_ip, password, device_code, address, longitude, latitude, device_company, device_type, server, time, other, owner_company, build_company)
        data = {
            "d_type": d_type,
            "camera_ip": camera_ip,
            "terminal_ip": terminal_ip,
            "password": password,
            "device_code": device_code,
            "address": address,
            "longitude": longitude,
            "latitude": latitude,
            "device_company": device_company,
            "device_type": device_type,
            "server": server,
            "time": time,
            "other": other,
            "owner_company": owner_company,
            "build_company": build_company,
            "road_code": road_code
        }
        # res = requests.post("http://111.230.246.188:8888/test/", data=data)
        res = requests.post("http://111.230.246.188:8899/test/", data=data)
        print(res)
        print(res.json())
        status = res.json()['status']

        if status == "success":
            ok += 1
        elif status == "fail":
            fail += 1

    print('失败', fail)
    print('添加成功', ok)

    #     address = rows[1]
    #     a = address.split('-')[-1]
    #     print(a)
    #     # res = re.findall(r'(?<=停在).*?(?=禁停路段)', str(rows[1]), re.S)
    #     # print(1, res)
    #     # if not res:
    #         # res = re.findall(r'(?<=在).*?(?=违反)', str(rows[1]), re.S)
    #         # outws.cell(i, 2).value = res[0]
    #         # continue
    #     outws.cell(i+1, 1).value = a
    #     # if not res:
    #     #     print(rows)
    #
    # saveExcel = "test2.xlsx"
    # outwb.save(saveExcel)  # 一定要记得保存


if __name__ == '__main__':
    read_excel()
